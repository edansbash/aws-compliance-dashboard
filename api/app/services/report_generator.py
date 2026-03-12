"""Report generation service for PDF dashboard and Excel exports."""

import io
import os
from datetime import datetime
from typing import Optional, List
from uuid import UUID

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.graphics.shapes import Drawing, Rect
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart

from app.models import Finding, Rule, Scan


# Color mappings for severity
SEVERITY_COLORS = {
    "CRITICAL": "dc2626",  # red-600
    "HIGH": "ea580c",      # orange-600
    "MEDIUM": "ca8a04",    # yellow-600
    "LOW": "2563eb",       # blue-600
}

SEVERITY_COLORS_RGB = {
    "CRITICAL": colors.HexColor("#dc2626"),
    "HIGH": colors.HexColor("#ea580c"),
    "MEDIUM": colors.HexColor("#ca8a04"),
    "LOW": colors.HexColor("#2563eb"),
}

STATUS_COLORS = {
    "PASS": "16a34a",      # green-600
    "FAIL": "dc2626",      # red-600
    "ERROR": "6b7280",     # gray-500
    "EXCEPTION": "7c3aed", # violet-600
}


async def get_findings_data(
    db: AsyncSession,
    scan_id: Optional[UUID] = None,
    account_ids: Optional[List[str]] = None,
    regions: Optional[List[str]] = None,
    severities: Optional[List[str]] = None,
    statuses: Optional[List[str]] = None,
) -> List[Finding]:
    """Fetch findings with optional filters."""
    conditions = []

    if scan_id:
        conditions.append(Finding.scan_id == scan_id)
    if account_ids:
        conditions.append(Finding.account_id.in_(account_ids))
    if regions:
        conditions.append(Finding.region.in_(regions))
    if statuses:
        conditions.append(Finding.status.in_(statuses))

    query = (
        select(Finding)
        .options(selectinload(Finding.rule))
        .order_by(Finding.created_at.desc())
    )

    if conditions:
        query = query.where(and_(*conditions))

    # Filter by severity requires join with Rule
    if severities:
        query = query.join(Rule).where(Rule.severity.in_(severities))

    result = await db.execute(query)
    return list(result.scalars().all())


async def get_summary_data(findings: List[Finding]) -> dict:
    """Calculate summary statistics from findings."""
    total = len(findings)
    by_status = {}
    by_severity = {}
    failing_by_severity = {}
    by_account = {}
    unique_resources = set()

    for f in findings:
        unique_resources.add(f.resource_id)
        by_status[f.status] = by_status.get(f.status, 0) + 1

        if f.rule:
            by_severity[f.rule.severity] = by_severity.get(f.rule.severity, 0) + 1
            # Track failed findings by severity
            if f.status == "FAIL":
                failing_by_severity[f.rule.severity] = failing_by_severity.get(f.rule.severity, 0) + 1

        if f.account_id not in by_account:
            by_account[f.account_id] = {"total": 0, "passing": 0, "failing": 0}
        by_account[f.account_id]["total"] += 1
        if f.status == "PASS":
            by_account[f.account_id]["passing"] += 1
        elif f.status == "FAIL":
            by_account[f.account_id]["failing"] += 1

    passing = by_status.get("PASS", 0) + by_status.get("EXCEPTION", 0)
    compliance_score = (passing / total * 100) if total > 0 else 100

    return {
        "total_findings": total,
        "total_resources": len(unique_resources),
        "compliance_score": round(compliance_score, 1),
        "by_status": by_status,
        "by_severity": by_severity,
        "failing_by_severity": failing_by_severity,
        "by_account": by_account,
    }


async def generate_excel_report(
    db: AsyncSession,
    scan_id: Optional[UUID] = None,
    filters: Optional[dict] = None,
) -> tuple[bytes, str]:
    """
    Generate Excel report of findings.

    Returns:
        Tuple of (file_bytes, filename)
    """
    filter_params = filters or {}
    findings = await get_findings_data(
        db,
        scan_id=scan_id,
        account_ids=filter_params.get("account_ids"),
        regions=filter_params.get("regions"),
        severities=filter_params.get("severities"),
        statuses=filter_params.get("statuses"),
    )

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary = await get_summary_data(findings)

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")

    # Summary header
    ws_summary["A1"] = "AWS Compliance Report - Full Export"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary.merge_cells("A1:D1")

    ws_summary["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws_summary.merge_cells("A2:D2")

    ws_summary["A3"] = f"Total Records Exported: {len(findings)}"
    ws_summary["A3"].font = Font(italic=True)
    ws_summary.merge_cells("A3:D3")

    # Summary stats
    ws_summary["A5"] = "Compliance Score"
    ws_summary["B5"] = f"{summary['compliance_score']}%"
    ws_summary["B5"].font = Font(bold=True, size=14)

    ws_summary["A6"] = "Total Findings"
    ws_summary["B6"] = summary["total_findings"]

    ws_summary["A7"] = "Total Resources"
    ws_summary["B7"] = summary["total_resources"]

    # By Status
    ws_summary["A9"] = "Findings by Status"
    ws_summary["A9"].font = Font(bold=True)
    row = 10
    for status, count in summary["by_status"].items():
        ws_summary[f"A{row}"] = status
        ws_summary[f"B{row}"] = count
        if status in STATUS_COLORS:
            ws_summary[f"A{row}"].fill = PatternFill(
                start_color=STATUS_COLORS[status],
                end_color=STATUS_COLORS[status],
                fill_type="solid"
            )
            ws_summary[f"A{row}"].font = Font(color="FFFFFF")
        row += 1

    # Failed Findings by Severity
    row += 1
    ws_summary[f"A{row}"] = "Failed Findings by Severity"
    ws_summary[f"A{row}"].font = Font(bold=True, color="dc2626")
    row += 1
    total_failed = 0
    for severity in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]:
        count = summary["failing_by_severity"].get(severity, 0)
        total_failed += count
        ws_summary[f"A{row}"] = severity
        ws_summary[f"B{row}"] = count
        if severity in SEVERITY_COLORS:
            ws_summary[f"A{row}"].fill = PatternFill(
                start_color=SEVERITY_COLORS[severity],
                end_color=SEVERITY_COLORS[severity],
                fill_type="solid"
            )
            ws_summary[f"A{row}"].font = Font(color="FFFFFF")
        # Highlight non-zero counts
        if count > 0:
            ws_summary[f"B{row}"].font = Font(bold=True, color="dc2626")
        row += 1
    # Total failed row
    ws_summary[f"A{row}"] = "TOTAL FAILED"
    ws_summary[f"B{row}"] = total_failed
    ws_summary[f"A{row}"].font = Font(bold=True)
    ws_summary[f"B{row}"].font = Font(bold=True, color="dc2626")
    row += 1

    # Findings sheet - Full export of all findings
    ws_findings = wb.create_sheet("Findings")

    headers = [
        "Finding ID", "Scan ID", "Resource ID", "Resource Name", "Resource Type",
        "Account ID", "Region", "Rule ID", "Rule Name", "Rule Description",
        "Severity", "Status", "Workflow Status", "Workflow Notes",
        "Details", "Created At"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws_findings.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Freeze header row
    ws_findings.freeze_panes = "A2"

    # Add ALL findings data
    for row_num, finding in enumerate(findings, 2):
        ws_findings.cell(row=row_num, column=1, value=str(finding.id))
        ws_findings.cell(row=row_num, column=2, value=str(finding.scan_id))
        ws_findings.cell(row=row_num, column=3, value=finding.resource_id)
        ws_findings.cell(row=row_num, column=4, value=finding.resource_name)
        ws_findings.cell(row=row_num, column=5, value=finding.resource_type)
        ws_findings.cell(row=row_num, column=6, value=finding.account_id)
        ws_findings.cell(row=row_num, column=7, value=finding.region)
        ws_findings.cell(row=row_num, column=8, value=finding.rule.rule_id if finding.rule else "")
        ws_findings.cell(row=row_num, column=9, value=finding.rule.name if finding.rule else "")
        ws_findings.cell(row=row_num, column=10, value=finding.rule.description if finding.rule else "")

        severity_cell = ws_findings.cell(row=row_num, column=11, value=finding.rule.severity if finding.rule else "")
        if finding.rule and finding.rule.severity in SEVERITY_COLORS:
            severity_cell.fill = PatternFill(
                start_color=SEVERITY_COLORS[finding.rule.severity],
                end_color=SEVERITY_COLORS[finding.rule.severity],
                fill_type="solid"
            )
            severity_cell.font = Font(color="FFFFFF")

        status_cell = ws_findings.cell(row=row_num, column=12, value=finding.status)
        if finding.status in STATUS_COLORS:
            status_cell.fill = PatternFill(
                start_color=STATUS_COLORS[finding.status],
                end_color=STATUS_COLORS[finding.status],
                fill_type="solid"
            )
            status_cell.font = Font(color="FFFFFF")

        ws_findings.cell(row=row_num, column=13, value=finding.workflow_status)
        ws_findings.cell(row=row_num, column=14, value=finding.workflow_notes or "")

        # Format details as readable string
        details_str = ""
        if finding.details:
            if isinstance(finding.details, dict):
                details_str = "; ".join(f"{k}: {v}" for k, v in finding.details.items())
            else:
                details_str = str(finding.details)
        ws_findings.cell(row=row_num, column=15, value=details_str)

        ws_findings.cell(row=row_num, column=16, value=finding.created_at.strftime("%Y-%m-%d %H:%M:%S"))

    # Auto-adjust column widths
    for ws in [ws_summary, ws_findings]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"compliance_findings_{timestamp}.xlsx"

    return output.getvalue(), filename


async def generate_pdf_dashboard(
    db: AsyncSession,
    scan_id: Optional[UUID] = None,
    filters: Optional[dict] = None,
) -> tuple[bytes, str]:
    """
    Generate PDF dashboard report.

    Returns:
        Tuple of (file_bytes, filename)
    """
    filter_params = filters or {}
    findings = await get_findings_data(
        db,
        scan_id=scan_id,
        account_ids=filter_params.get("account_ids"),
        regions=filter_params.get("regions"),
        severities=filter_params.get("severities"),
        statuses=filter_params.get("statuses"),
    )

    summary = await get_summary_data(findings)

    # Create PDF in memory
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=letter,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch,
    )

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.HexColor("#1e3a5f"),
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=20,
        spaceAfter=10,
        textColor=colors.HexColor("#1e3a5f"),
    )

    elements = []

    # Title
    elements.append(Paragraph("AWS Compliance Dashboard Report", title_style))
    elements.append(Paragraph(
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 20))

    # Compliance Score Card
    score_color = colors.green if summary["compliance_score"] >= 80 else (
        colors.orange if summary["compliance_score"] >= 60 else colors.red
    )

    score_data = [
        ["Compliance Score", "Total Findings", "Total Resources"],
        [
            f"{summary['compliance_score']}%",
            str(summary["total_findings"]),
            str(summary["total_resources"]),
        ]
    ]

    score_table = Table(score_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
    score_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, 1), 18),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 1), (-1, 1), 15),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 15),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#e5e7eb")),
    ]))
    elements.append(score_table)
    elements.append(Spacer(1, 30))

    # Findings by Status
    elements.append(Paragraph("Findings by Status", heading_style))

    status_data = [["Status", "Count"]]
    for status in ["PASS", "FAIL", "ERROR", "EXCEPTION"]:
        count = summary["by_status"].get(status, 0)
        status_data.append([status, str(count)])

    status_table = Table(status_data, colWidths=[3*inch, 2*inch])
    status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#e5e7eb")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
    ]))
    elements.append(status_table)
    elements.append(Spacer(1, 20))

    # Findings by Severity
    elements.append(Paragraph("Findings by Severity", heading_style))

    severity_data = [["Severity", "Count"]]
    for severity in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]:
        count = summary["by_severity"].get(severity, 0)
        severity_data.append([severity, str(count)])

    severity_table = Table(severity_data, colWidths=[3*inch, 2*inch])
    severity_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#e5e7eb")),
    ]

    # Add severity colors
    for i, severity in enumerate(["CRITICAL", "HIGH", "MEDIUM", "LOW"], 1):
        severity_style.append(('BACKGROUND', (0, i), (0, i), SEVERITY_COLORS_RGB.get(severity, colors.white)))
        severity_style.append(('TEXTCOLOR', (0, i), (0, i), colors.white))

    severity_table.setStyle(TableStyle(severity_style))
    elements.append(severity_table)
    elements.append(Spacer(1, 20))

    # Findings by Account
    if summary["by_account"]:
        elements.append(Paragraph("Findings by Account", heading_style))

        account_data = [["Account ID", "Total", "Passing", "Failing"]]
        for account_id, data in summary["by_account"].items():
            account_data.append([
                account_id,
                str(data["total"]),
                str(data["passing"]),
                str(data["failing"]),
            ])

        account_table = Table(account_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        account_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#e5e7eb")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
        ]))
        elements.append(account_table)
        elements.append(Spacer(1, 20))

    # Top Failing Findings (limited to 10)
    failing_findings = [f for f in findings if f.status == "FAIL"][:10]

    if failing_findings:
        elements.append(Paragraph("Top Non-Compliant Resources", heading_style))

        findings_data = [["Resource", "Rule", "Severity", "Account"]]
        for f in failing_findings:
            findings_data.append([
                f.resource_name[:30] + "..." if len(f.resource_name) > 30 else f.resource_name,
                f.rule.name[:25] + "..." if f.rule and len(f.rule.name) > 25 else (f.rule.name if f.rule else ""),
                f.rule.severity if f.rule else "",
                f.account_id,
            ])

        findings_table = Table(findings_data, colWidths=[2.2*inch, 2.2*inch, 1.3*inch, 1.3*inch])
        findings_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#e5e7eb")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
        ]))
        elements.append(findings_table)

    # Build PDF
    doc.build(elements)
    output.seek(0)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"compliance_dashboard_{timestamp}.pdf"

    return output.getvalue(), filename
